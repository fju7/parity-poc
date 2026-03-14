#!/usr/bin/env python3
"""Load Medicare Part B ASP (Average Sales Price) data into Supabase.

Downloads the CMS ASP pricing ZIP file, extracts the pricing CSV/TXT,
parses it, and upserts all records into the pharmacy_asp table.

Usage:
    SUPABASE_URL=... SUPABASE_SERVICE_ROLE_KEY=... python3 backend/scripts/load_asp.py

Requires: pip3 install requests supabase openpyxl --break-system-packages
"""

import io
import os
import re
import sys
import zipfile
from typing import Optional

import requests

# ---------------------------------------------------------------------------
# CMS ASP download URLs — try in order
# ---------------------------------------------------------------------------

ASP_URLS = [
    ("https://www.cms.gov/files/zip/april-2026-medicare-part-b-payment-limit-files-03-10-2026-preliminary-file.zip", "2026-Q2-preliminary"),
    ("https://www.cms.gov/files/zip/january-2026-medicare-part-b-payment-limit-files-03-10-2026-final-file.zip", "2026-Q1"),
]


def get_supabase():
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set.")
        sys.exit(1)
    from supabase import create_client
    return create_client(url, key)


def download_asp_zip() -> tuple:
    """Try each CMS URL until one works. Returns (ZIP bytes, effective_quarter)."""
    for url, quarter in ASP_URLS:
        print(f"  Trying: {url}")
        try:
            resp = requests.get(url, timeout=120, stream=True)
            if resp.status_code == 200:
                data = resp.content
                print(f"  Downloaded {len(data):,} bytes from {url}")
                print(f"  Effective quarter: {quarter}")
                return data, quarter
            else:
                print(f"  HTTP {resp.status_code} — trying next URL...")
        except Exception as exc:
            print(f"  Failed: {exc} — trying next URL...")

    print("ERROR: All ASP download URLs failed.")
    sys.exit(1)


def find_pricing_file(z: zipfile.ZipFile) -> str:
    """Find the main ASP pricing file inside the ZIP.

    Looks for files with 'asp' and 'pricing' in the name, or files
    ending in .txt/.csv that contain HCPCS-like data.
    """
    names = z.namelist()
    print(f"  ZIP contains {len(names)} files:")
    for n in names:
        print(f"    {n} ({z.getinfo(n).file_size:,} bytes)")

    # Priority 1: file with "asp" and "pricing" in name
    for n in names:
        nl = n.lower()
        if ("asp" in nl and "pricing" in nl) and (nl.endswith(".txt") or nl.endswith(".csv")):
            return n

    # Priority 2: file with "asp" in name, .txt or .csv
    for n in names:
        nl = n.lower()
        if "asp" in nl and (nl.endswith(".txt") or nl.endswith(".csv")):
            return n

    # Priority 3: file with "payment" in name
    for n in names:
        nl = n.lower()
        if "payment" in nl and (nl.endswith(".txt") or nl.endswith(".csv") or nl.endswith(".xlsx")):
            return n

    # Priority 4: any .xlsx file (CMS sometimes uses Excel)
    for n in names:
        if n.lower().endswith(".xlsx"):
            return n

    # Priority 5: largest .txt or .csv file
    txt_csv = [n for n in names if n.lower().endswith((".txt", ".csv"))]
    if txt_csv:
        return max(txt_csv, key=lambda n: z.getinfo(n).file_size)

    print("ERROR: Could not find ASP pricing file in ZIP.")
    print(f"  Files found: {names}")
    sys.exit(1)


def parse_asp_file(z: zipfile.ZipFile, filename: str) -> list[dict]:
    """Parse the ASP pricing file and return list of records.

    CMS ASP file structure (confirmed):
      Rows 0-7:  header/notes — skip
      Row 8:     column headers — "HCPCS Code", "Short Description",
                 "HCPCS Code Dosage", "Payment Limit", ...
      Rows 9+:   data rows

    Fixed column indices:
      Col 0 = hcpcs_code
      Col 1 = short_description
      Col 2 = dosage
      Col 3 = payment_limit (numeric)
    """
    print(f"  Parsing: {filename}")

    if filename.lower().endswith(".xlsx"):
        return _parse_xlsx(z, filename)
    else:
        return _parse_text(z, filename)


def _parse_xlsx(z: zipfile.ZipFile, filename: str) -> list[dict]:
    """Parse an Excel ASP file using fixed column positions."""
    import openpyxl

    data = z.read(filename)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    rows_out = []

    # Find header row: first cell is exactly "HCPCS Code"
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and str(row[0] or "").strip().lower() == "hcpcs code":
            header_row = i
            print(f"  Found header at Excel row {i}")
            break

    if header_row is None:
        print("  WARNING: Could not find 'HCPCS Code' header row in Excel file")
        wb.close()
        return []

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= header_row:
            continue
        if not row or len(row) < 4:
            continue

        record = _extract_fixed_columns(list(row))
        if record:
            rows_out.append(record)

    wb.close()
    return rows_out


def _parse_text(z: zipfile.ZipFile, filename: str) -> list[dict]:
    """Parse a text/CSV ASP file using fixed column positions."""
    data = z.read(filename).decode("utf-8", errors="replace")
    lines = data.strip().split("\n")
    rows_out = []

    # Detect delimiter from first non-empty line
    sample_line = lines[0] if lines else ""
    if "\t" in sample_line:
        delimiter = "\t"
    elif "," in sample_line:
        delimiter = ","
    else:
        delimiter = ","  # default to comma

    # Find header row: first cell is exactly "HCPCS Code"
    header_idx = None
    for i, line in enumerate(lines):
        cells = [c.strip().strip('"').strip() for c in line.split(delimiter)]
        if cells and cells[0].lower() == "hcpcs code":
            header_idx = i
            print(f"  Found header at line {i}")
            break

    if header_idx is None:
        print("  WARNING: Could not find 'HCPCS Code' header row in text file")
        return []

    # Parse data lines after header
    for line in lines[header_idx + 1:]:
        line = line.strip()
        if not line:
            continue

        cells = [c.strip().strip('"').strip() for c in line.split(delimiter)]
        record = _extract_fixed_columns(cells)
        if record:
            rows_out.append(record)

    return rows_out


def _extract_fixed_columns(cells: list) -> Optional[dict]:
    """Extract an ASP record using fixed column positions.

    Col 0 = hcpcs_code (must match pattern: letter + 4 digits)
    Col 1 = short_description
    Col 2 = dosage
    Col 3 = payment_limit (numeric, required)
    """
    if not cells or len(cells) < 4:
        return None

    # Column 0: HCPCS code
    hcpcs = str(cells[0] or "").strip().upper()
    if not re.match(r"^[A-Z]\d{4}$", hcpcs):
        return None

    # Column 3: payment limit (required, must be numeric)
    limit_str = str(cells[3] or "").strip().replace("$", "").replace(",", "")
    if not limit_str:
        return None
    try:
        payment_limit = float(limit_str)
    except (ValueError, TypeError):
        return None

    # Column 1: short description
    short_desc = str(cells[1] or "").strip()

    # Column 2: dosage
    dosage = str(cells[2] or "").strip()

    return {
        "hcpcs_code": hcpcs,
        "short_description": short_desc[:200] if short_desc else None,
        "dosage": dosage[:100] if dosage else None,
        "payment_limit": payment_limit,
    }



def upsert_to_supabase(sb, rows: list[dict], quarter: str) -> int:
    """Clear existing data and insert all ASP records."""
    print(f"  Clearing existing pharmacy_asp data...")
    try:
        sb.table("pharmacy_asp").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
    except Exception as exc:
        print(f"  Warning: could not clear table (may be empty): {exc}")

    batch_size = 200
    total = len(rows)
    inserted = 0

    for i in range(0, total, batch_size):
        batch = rows[i:i + batch_size]
        records = [
            {
                "hcpcs_code": r["hcpcs_code"],
                "short_description": r.get("short_description"),
                "dosage": r.get("dosage"),
                "payment_limit": r["payment_limit"],
                "effective_quarter": quarter,
            }
            for r in batch
        ]
        try:
            sb.table("pharmacy_asp").insert(records).execute()
            inserted += len(batch)
            print(f"  Inserted {inserted}/{total} rows...")
        except Exception as exc:
            print(f"  Error inserting batch at offset {i}: {exc}")
            continue

    return inserted


def main():
    print("=== Medicare Part B ASP Loader ===")

    print("1. Downloading ASP pricing file from CMS...")
    zip_bytes, quarter = download_asp_zip()

    print("2. Extracting pricing data from ZIP...")
    z = zipfile.ZipFile(io.BytesIO(zip_bytes))
    pricing_file = find_pricing_file(z)
    print(f"  Selected: {pricing_file}")

    rows = parse_asp_file(z, pricing_file)
    print(f"  Parsed {len(rows)} ASP records.")

    if not rows:
        print("  ERROR: No records parsed. Check the file format.")
        sys.exit(1)

    # Show sample
    print("  Sample records:")
    for r in rows[:5]:
        print(f"    {r['hcpcs_code']} | {r.get('short_description', '')[:40]} | "
              f"${r['payment_limit']:.4f} | {r.get('dosage', '')}")
    print(f"  Effective quarter: {quarter}")

    print("3. Connecting to Supabase...")
    sb = get_supabase()

    print("4. Upserting to pharmacy_asp table...")
    count = upsert_to_supabase(sb, rows, quarter)
    print(f"   Done! {count} ASP records loaded.")

    # Summary stats
    unique_hcpcs = len(set(r["hcpcs_code"] for r in rows))
    print(f"   Unique HCPCS codes: {unique_hcpcs}")

    # Check for common drugs
    known_drugs = {
        "J0185": "Humira (adalimumab)",
        "J3490": "Unclassified drugs",
        "J7170": "Ozempic/semaglutide",
        "Q5131": "Adalimumab-adbm (biosimilar)",
        "J0881": "Darbepoetin alfa",
        "J2505": "Pegfilgrastim",
        "J9035": "Bevacizumab",
        "J9271": "Pembrolizumab (Keytruda)",
    }
    print("   Known drug check:")
    rows_by_hcpcs = {r["hcpcs_code"]: r for r in rows}
    for code, name in known_drugs.items():
        if code in rows_by_hcpcs:
            r = rows_by_hcpcs[code]
            print(f"     {code} ({name}): FOUND — desc=\"{r.get('short_description', '')}\" limit=${r['payment_limit']:.4f}")
        else:
            print(f"     {code} ({name}): not found")

    # Validate J0185 specifically — should have meaningful description and high payment limit
    if "J0185" in rows_by_hcpcs:
        j0185 = rows_by_hcpcs["J0185"]
        desc = j0185.get("short_description", "") or ""
        limit = j0185.get("payment_limit", 0) or 0
        if len(desc) <= 5:
            print(f"   WARNING: J0185 short_description looks truncated: \"{desc}\"")
            print("   The CSV parser may not be extracting descriptions correctly.")
        if limit < 100:
            print(f"   WARNING: J0185 payment_limit is ${limit:.4f} — expected > $100")
            print("   The CSV parser may be reading the wrong column for payment_limit.")


if __name__ == "__main__":
    main()
