"""Generate realistic test files for CivicScale platform testing.

Creates 5 test files:
1. Provider 835 EDI remittance (Chesapeake Family Medicine, 2 payers)
2. Provider 837P EDI professional claims
3. Provider contract fee schedule (BlueCross Maryland, Excel)
4. Employer 835 EDI remittance (Maryland Precision Manufacturing)
5. Employer SBC PDF (Maryland Precision Manufacturing PPO 2024)

All data is fictional but structurally valid for testing EDI parsing,
contract integrity analysis, coding analysis, claims check, and scorecard.
"""

import os
import random
import string
from datetime import datetime, timedelta

random.seed(2024)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "test_data")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ═══════════════════════════════════════════════════════════════════
# Constants
# ═══════════════════════════════════════════════════════════════════

PRACTICE_NAME = "Chesapeake Family Medicine"
PRACTICE_NPI = "1234567890"
PRACTICE_TIN = "520123456"
PRACTICE_ZIP = "21201"

PAYER_BLUECROSS = "BlueCross Maryland"
PAYER_BLUECROSS_ID = "BC1234"
PAYER_AETNA = "Aetna Mid-Atlantic"
PAYER_AETNA_ID = "AE5678"

EMPLOYER_NAME = "Maryland Precision Manufacturing Inc."
EMPLOYER_TIN = "521234567"
EMPLOYER_PLAN = "Maryland Manufacturing Health Plan"
EMPLOYER_PLAN_ID = "MPM2024"

# Maryland Medicare rates (2024, locality 12)
MEDICARE_RATES = {
    "99213": 97, "99214": 138, "99215": 196,
    "99203": 121, "99204": 168, "99205": 232,
    "99396": 166, "99397": 187,
    "99441": 48,
    "93000": 30, "36415": 13, "85025": 11,
    "80053": 18, "82947": 9, "83036": 14,
    "G0439": 175,
    # Additional Family Medicine codes for contract
    "99201": 48, "99202": 76, "99211": 26, "99212": 59,
    "99391": 130, "99392": 140, "99393": 145, "99394": 155, "99395": 160,
    "99381": 120, "99382": 130, "99383": 135, "99384": 145, "99385": 150,
    "87804": 16, "87880": 20, "71046": 32,
    "90471": 30, "90472": 18, "90658": 20, "90686": 25,
    "81002": 5, "81003": 5, "36416": 5,
    "10060": 155, "11102": 145, "17110": 55,
    "20610": 70,
    # Employer-specific codes
    "99243": 145, "99244": 210,
    "27447": 1800, "66984": 1200, "43239": 450, "23472": 1400,
    "72148": 240, "73721": 220,
    "99283": 150, "99284": 260,
    "90837": 120, "90834": 95,
}

# Contracted rates at 118% of Medicare
CONTRACTED_RATES = {k: round(v * 1.18, 2) for k, v in MEDICARE_RATES.items()}

# Provider 835 CPT mix for Family Medicine
PROVIDER_CPTS = [
    "99213", "99213", "99213", "99213", "99213",  # high volume
    "99214", "99214", "99214", "99214",
    "99215", "99215",
    "99203", "99203",
    "99204",
    "99396", "99396", "99396",
    "99397", "99397",
    "99441",
    "93000", "93000",
    "36415", "36415", "36415",
    "85025", "85025",
    "80053", "80053",
    "82947",
    "G0439", "G0439",
]

# ICD-10 codes by CPT
ICD_MAP = {
    "99213": ["I10", "E11.9", "J06.9", "M54.5"],
    "99214": ["I10", "E11.9", "E78.5", "I25.10"],
    "99215": ["E11.9", "I10", "E78.5"],
    "99203": ["Z00.00", "I10", "R05.9"],
    "99204": ["Z00.01", "E11.65"],
    "99396": ["Z00.00"],
    "99397": ["Z00.00"],
    "99441": ["I10", "E11.9"],
    "93000": ["I10", "R00.0"],
    "36415": ["Z01.89"],
    "85025": ["D64.9", "Z01.89"],
    "80053": ["E11.9", "Z01.89"],
    "82947": ["E11.9"],
    "G0439": ["Z00.00"],
}

# CARC denial codes
DENIAL_CODES = [
    ("CO", "4", "The procedure code is inconsistent with the modifier used"),
    ("CO", "97", "The benefit for this service is included in the payment for another service"),
    ("PR", "1", "Deductible amount"),
    ("CO", "45", "Charge exceeds fee schedule/maximum allowable"),
]

# Fictional patient names
PATIENTS = [
    ("JOHNSON", "MICHAEL", "M", "19750315"),
    ("WILLIAMS", "SARAH", "F", "19820622"),
    ("BROWN", "DAVID", "M", "19680910"),
    ("GARCIA", "MARIA", "F", "19901204"),
    ("DAVIS", "ROBERT", "M", "19550718"),
    ("MILLER", "JENNIFER", "F", "19780503"),
    ("WILSON", "THOMAS", "M", "19850129"),
    ("MOORE", "LISA", "F", "19720815"),
    ("TAYLOR", "JAMES", "M", "19600427"),
    ("ANDERSON", "PATRICIA", "F", "19880711"),
    ("THOMAS", "CHRISTOPHER", "M", "19950216"),
    ("JACKSON", "AMANDA", "F", "19700930"),
    ("WHITE", "WILLIAM", "M", "19830105"),
    ("HARRIS", "ELIZABETH", "F", "19650820"),
    ("MARTIN", "DANIEL", "M", "19920614"),
]


def _random_claim_id():
    return "".join(random.choices(string.digits, k=12))


def _random_date_2024(month_start=1, month_end=6):
    month = random.randint(month_start, month_end)
    day = random.randint(1, 28)
    return f"2024{month:02d}{day:02d}"


def _format_date_dash(edi_date):
    return f"{edi_date[:4]}-{edi_date[4:6]}-{edi_date[6:8]}"


# ═══════════════════════════════════════════════════════════════════
# File 1: Provider 835 EDI
# ═══════════════════════════════════════════════════════════════════

def generate_provider_835():
    """Generate 835 remittance for Chesapeake Family Medicine."""
    lines = []
    claim_data = []  # Store for 837 generation
    seg_count = 0

    def add(seg):
        nonlocal seg_count
        lines.append(seg)
        seg_count += 1

    now = datetime.now().strftime("%y%m%d")
    time_now = datetime.now().strftime("%H%M")

    # ISA header
    add(f"ISA*00*          *00*          *ZZ*{PAYER_BLUECROSS_ID:<15}*ZZ*{PRACTICE_TIN:<15}*{now}*{time_now}*^*00501*000000001*0*T*:~")
    add(f"GS*HP*{PAYER_BLUECROSS_ID}*{PRACTICE_TIN}*20240630*{time_now}*1*X*005010X221A1~")
    add("ST*835*0001~")
    add(f"BPR*I*0*C*ACH*CCP*01*021000021*DA*123456789*{PAYER_BLUECROSS_ID}**01*021000021*DA*987654321*20240715~")
    add("TRN*1*TRACE00001*1021000021~")
    add(f"DTM*405*20240630~")
    add(f"N1*PR*{PAYER_BLUECROSS}~")
    add(f"N3*PO BOX 12345~")
    add("N4*BALTIMORE*MD*21202~")
    add(f"N1*PE*{PRACTICE_NAME}*XX*{PRACTICE_NPI}~")
    add(f"N3*100 HARBOR VIEW DR~")
    add(f"N4*BALTIMORE*MD*{PRACTICE_ZIP}~")

    claim_num = 0

    def add_claims(payer_name, payer_id, count):
        nonlocal claim_num
        for _ in range(count):
            patient = random.choice(PATIENTS)
            cpt = random.choice(PROVIDER_CPTS)
            service_date = _random_date_2024()
            contracted = CONTRACTED_RATES.get(cpt, 100)
            billed = round(contracted * 1.4, 2)
            claim_id = _random_claim_id()
            member_id = f"MBR{random.randint(100000, 999999)}"

            # Determine payment outcome
            roll = random.random()
            if roll < 0.70:
                # Paid at contracted rate
                paid = contracted
                adj_code = None
                adj_amt = round(billed - paid, 2)
            elif roll < 0.85:
                # Underpaid 10-20%
                underpay_pct = random.uniform(0.10, 0.20)
                paid = round(contracted * (1 - underpay_pct), 2)
                adj_code = None
                adj_amt = round(billed - paid, 2)
            elif roll < 0.95:
                # Denied
                paid = 0
                denial = random.choice(DENIAL_CODES)
                adj_code = denial
                adj_amt = billed
            else:
                # Partially paid (deductible/copay split)
                copay = random.choice([25, 30, 40, 50])
                paid = round(contracted - copay, 2)
                adj_code = ("PR", "2", "Coinsurance amount")
                adj_amt = round(billed - paid, 2)

            icd = random.choice(ICD_MAP.get(cpt, ["Z00.00"]))

            # CLP segment (claim level)
            status = "1" if paid > 0 else "4"  # 1=processed, 4=denied
            add(f"CLP*{claim_id}*{status}*{billed}*{paid}*{round(adj_amt, 2)}*MC*CLAIM{claim_num:06d}~")
            add(f"NM1*QC*1*{patient[0]}*{patient[1]}****MI*{member_id}~")

            # SVC segment (service line)
            add(f"SVC*HC:{cpt}*{billed}*{paid}**1~")
            add(f"DTM*472*{service_date}~")

            if adj_code:
                add(f"CAS*{adj_code[0]}*{adj_code[1]}*{round(adj_amt, 2)}~")
            elif adj_amt > 0:
                add(f"CAS*CO*45*{round(adj_amt, 2)}~")

            claim_data.append({
                "claim_id": claim_id,
                "patient": patient,
                "member_id": member_id,
                "cpt": cpt,
                "icd": icd,
                "service_date": service_date,
                "billed": billed,
                "paid": paid,
                "payer_name": payer_name,
                "payer_id": payer_id,
            })
            claim_num += 1

    # BlueCross claims (60%)
    add_claims(PAYER_BLUECROSS, PAYER_BLUECROSS_ID, 55)

    # Switch to Aetna payer header
    add(f"N1*PR*{PAYER_AETNA}~")
    add("N3*PO BOX 67890~")
    add("N4*COLUMBIA*MD*21046~")

    # Aetna claims (40%)
    add_claims(PAYER_AETNA, PAYER_AETNA_ID, 35)

    # Trailer
    add(f"SE*{seg_count + 1}*0001~")
    add("GE*1*1~")
    add("IEA*1*000000001~")

    output = "\n".join(lines)
    path = os.path.join(OUTPUT_DIR, "chesapeake_835_2024.edi")
    with open(path, "w") as f:
        f.write(output)

    print(f"  [1] Provider 835: {path} ({len(claim_data)} claims, {os.path.getsize(path):,} bytes)")
    return claim_data


# ═══════════════════════════════════════════════════════════════════
# File 2: Provider 837P EDI
# ═══════════════════════════════════════════════════════════════════

def generate_provider_837(claim_data):
    """Generate 837P professional claims from the same claims as the 835."""
    lines = []
    seg_count = 0

    def add(seg):
        nonlocal seg_count
        lines.append(seg)
        seg_count += 1

    now = datetime.now().strftime("%y%m%d")
    time_now = datetime.now().strftime("%H%M")

    # ISA/GS/ST envelope
    add(f"ISA*00*          *00*          *ZZ*{PRACTICE_TIN:<15}*ZZ*{PAYER_BLUECROSS_ID:<15}*{now}*{time_now}*^*00501*000000002*0*T*:~")
    add(f"GS*HC*{PRACTICE_TIN}*{PAYER_BLUECROSS_ID}*20240630*{time_now}*2*X*005010X222A1~")
    add("ST*837*0001~")
    add("BHT*0019*00*BATCH001*20240630*1200*CH~")

    # Submitter
    add(f"NM1*41*2*{PRACTICE_NAME}*****46*{PRACTICE_TIN}~")
    add("PER*IC*BILLING DEPT*TE*4105551234~")

    # Receiver
    add(f"NM1*40*2*{PAYER_BLUECROSS}*****46*{PAYER_BLUECROSS_ID}~")

    # Billing provider
    add(f"HL*1**20*1~")
    add(f"NM1*85*2*{PRACTICE_NAME}*****XX*{PRACTICE_NPI}~")
    add(f"N3*100 HARBOR VIEW DR~")
    add(f"N4*BALTIMORE*MD*{PRACTICE_ZIP}~")
    add(f"REF*EI*{PRACTICE_TIN}~")

    hl_counter = 2
    for i, claim in enumerate(claim_data):
        # Subscriber HL
        add(f"HL*{hl_counter}*1*22*0~")
        add("SBR*P*18*******MC~")
        add(f"NM1*IL*1*{claim['patient'][0]}*{claim['patient'][1]}****MI*{claim['member_id']}~")
        add(f"N3*{random.randint(100, 9999)} MAIN ST~")
        add(f"N4*BALTIMORE*MD*{PRACTICE_ZIP}~")
        add(f"DMG*D8*{claim['patient'][3]}*{claim['patient'][2]}~")

        # Payer
        add(f"NM1*PR*2*{claim['payer_name']}*****PI*{claim['payer_id']}~")

        # Claim
        add(f"CLM*{claim['claim_id']}*{claim['billed']}***11:B:1*Y*A*Y*Y~")

        # Diagnosis
        add(f"HI*ABK:{claim['icd']}~")

        # Service line
        add(f"LX*1~")
        add(f"SV1*HC:{claim['cpt']}*{claim['billed']}*UN*1***1~")
        add(f"DTP*472*D8*{claim['service_date']}~")

        hl_counter += 1

    # Trailer
    add(f"SE*{seg_count + 1}*0001~")
    add("GE*1*2~")
    add("IEA*1*000000002~")

    output = "\n".join(lines)
    path = os.path.join(OUTPUT_DIR, "chesapeake_837_2024.edi")
    with open(path, "w") as f:
        f.write(output)

    print(f"  [2] Provider 837P: {path} ({len(claim_data)} claims, {os.path.getsize(path):,} bytes)")


# ═══════════════════════════════════════════════════════════════════
# File 3: Provider Contract Fee Schedule (Excel)
# ═══════════════════════════════════════════════════════════════════

def generate_contract_excel():
    """Generate BlueCross Maryland contract rates Excel matching the app template."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contract Rates"

    # Header rows matching the template format
    ws.merge_cells("A1:L1")
    ws["A1"] = "PARITY PROVIDER — Contract Rates Template"
    ws["A1"].font = Font(bold=True, size=14)

    ws.merge_cells("A2:L2")
    ws["A2"] = "Enter your contracted rate for each CPT code and payer.  Leave blank if not applicable.  Save and upload to Parity Provider."
    ws["A2"].font = Font(size=10, color="666666")

    # Column headers (row 4)
    headers = ["CPT Code", "Description", "Category", "BlueCross Maryland", "Aetna Mid-Atlantic"]
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col >= 4 else "left")

    # CPT codes with descriptions and categories
    cpt_rows = [
        ("99211", "Office visit, established, minimal", "E&M — Office"),
        ("99212", "Office visit, established, straightforward", "E&M — Office"),
        ("99213", "Office visit, established, low complexity", "E&M — Office"),
        ("99214", "Office visit, established, moderate complexity", "E&M — Office"),
        ("99215", "Office visit, established, high complexity", "E&M — Office"),
        ("99201", "Office visit, new patient, straightforward", "E&M — Office"),
        ("99202", "Office visit, new patient, straightforward", "E&M — Office"),
        ("99203", "Office visit, new patient, low complexity", "E&M — Office"),
        ("99204", "Office visit, new patient, moderate complexity", "E&M — Office"),
        ("99205", "Office visit, new patient, high complexity", "E&M — Office"),
        ("99381", "Preventive visit, new, infant", "Preventive"),
        ("99382", "Preventive visit, new, 1-4 years", "Preventive"),
        ("99383", "Preventive visit, new, 5-11 years", "Preventive"),
        ("99384", "Preventive visit, new, 12-17 years", "Preventive"),
        ("99385", "Preventive visit, new, 18-39 years", "Preventive"),
        ("99391", "Preventive visit, established, infant", "Preventive"),
        ("99392", "Preventive visit, established, 1-4 years", "Preventive"),
        ("99393", "Preventive visit, established, 5-11 years", "Preventive"),
        ("99394", "Preventive visit, established, 12-17 years", "Preventive"),
        ("99395", "Preventive visit, established, 18-39 years", "Preventive"),
        ("99396", "Preventive visit, established, 40-64 years", "Preventive"),
        ("99397", "Preventive visit, established, 65+ years", "Preventive"),
        ("99441", "Telephone E&M, 5-10 minutes", "Telehealth"),
        ("G0439", "Annual wellness visit, subsequent", "Preventive"),
        ("93000", "Electrocardiogram (ECG), 12-lead", "Diagnostic"),
        ("36415", "Venipuncture (blood draw)", "Lab"),
        ("36416", "Capillary blood collection", "Lab"),
        ("85025", "CBC with differential", "Lab"),
        ("80053", "Comprehensive metabolic panel", "Lab"),
        ("82947", "Glucose, quantitative", "Lab"),
        ("83036", "Hemoglobin A1c", "Lab"),
        ("81002", "Urinalysis, non-automated", "Lab"),
        ("87804", "Influenza assay with optical", "Lab"),
        ("87880", "Strep A test, rapid", "Lab"),
        ("71046", "Chest X-ray, 2 views", "Imaging"),
        ("90471", "Immunization administration", "Immunization"),
        ("90472", "Immunization admin, each additional", "Immunization"),
        ("10060", "Incision and drainage, simple", "Procedure"),
        ("11102", "Tangential biopsy of skin", "Procedure"),
        ("17110", "Destruction of warts, up to 14", "Procedure"),
        ("20610", "Arthrocentesis, major joint", "Procedure"),
    ]

    for i, (cpt, desc, cat) in enumerate(cpt_rows):
        row = i + 5
        ws.cell(row=row, column=1, value=cpt)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=cat)

        # BlueCross rate at 118% Medicare
        medicare = MEDICARE_RATES.get(cpt, 50)
        bc_rate = round(medicare * 1.18, 2)
        ws.cell(row=row, column=4, value=bc_rate).number_format = "#,##0.00"

        # Aetna rate at 115% Medicare (slightly lower)
        aetna_rate = round(medicare * 1.15, 2)
        ws.cell(row=row, column=5, value=aetna_rate).number_format = "#,##0.00"

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22

    path = os.path.join(OUTPUT_DIR, "chesapeake_contract_bluecross.xlsx")
    wb.save(path)
    print(f"  [3] Contract Excel: {path} ({len(cpt_rows)} CPT codes, {os.path.getsize(path):,} bytes)")


# ═══════════════════════════════════════════════════════════════════
# File 4: Employer 835 EDI
# ═══════════════════════════════════════════════════════════════════

def generate_employer_835():
    """Generate employer 835 for Maryland Precision Manufacturing."""
    lines = []
    seg_count = 0

    def add(seg):
        nonlocal seg_count
        lines.append(seg)
        seg_count += 1

    now = datetime.now().strftime("%y%m%d")
    time_now = datetime.now().strftime("%H%M")

    # Envelope
    add(f"ISA*00*          *00*          *ZZ*{EMPLOYER_PLAN_ID:<15}*ZZ*{EMPLOYER_TIN:<15}*{now}*{time_now}*^*00501*000000003*0*T*:~")
    add(f"GS*HP*{EMPLOYER_PLAN_ID}*{EMPLOYER_TIN}*20241231*{time_now}*3*X*005010X221A1~")
    add("ST*835*0001~")
    add(f"BPR*I*0*C*ACH*CCP*01*021000021*DA*111222333*{EMPLOYER_PLAN_ID}**01*021000021*DA*444555666*20250115~")
    add("TRN*1*EMPTRACE001*1021000021~")
    add("DTM*405*20241231~")
    add(f"N1*PR*{EMPLOYER_PLAN}~")
    add("N3*PO BOX 99999~")
    add("N4*ANNAPOLIS*MD*21401~")
    add(f"N1*PE*VARIOUS PROVIDERS*XX*9999999999~")

    # Employee and dependent names
    members = [
        ("CHEN", "ROBERT", "M", "EMP001", "19780315"),
        ("CHEN", "LINDA", "F", "DEP001A", "19800622"),
        ("CHEN", "EMMA", "F", "DEP001B", "20100415"),
        ("PATEL", "VIKRAM", "M", "EMP002", "19650910"),
        ("PATEL", "PRIYA", "F", "DEP002A", "19681204"),
        ("GONZALEZ", "CARLOS", "M", "EMP003", "19850718"),
        ("GONZALEZ", "MARIA", "F", "DEP003A", "19870503"),
        ("SMITH", "JOHN", "M", "EMP004", "19720129"),
        ("SMITH", "KAREN", "F", "DEP004A", "19740815"),
        ("SMITH", "TYLER", "M", "DEP004B", "20050427"),
        ("NGUYEN", "TRAN", "M", "EMP005", "19600711"),
        ("JOHNSON", "MARCUS", "M", "EMP006", "19880216"),
        ("WILLIAMS", "DEANNA", "F", "EMP007", "19750930"),
        ("BROWN", "KEVIN", "M", "EMP008", "19830105"),
        ("DAVIS", "ANGELA", "F", "EMP009", "19700820"),
        ("MILLER", "SCOTT", "M", "EMP010", "19920614"),
        ("WILSON", "JANET", "F", "EMP011", "19680311"),
        ("MOORE", "BRIAN", "M", "EMP012", "19850929"),
        ("TAYLOR", "KRISTIN", "F", "EMP013", "19790517"),
        ("ANDERSON", "PAUL", "M", "EMP014", "19550808"),
    ]

    # Employer CPT mix with realistic weights
    employer_cpts = {
        # Primary care
        "99213": 35, "99214": 25,
        # Specialist visits
        "99243": 8, "99244": 5,
        # Procedures (high cost)
        "27447": 2, "66984": 2, "43239": 3, "23472": 1,
        # Labs
        "80053": 15, "85025": 12, "82947": 8, "83036": 10,
        # Imaging
        "71046": 8, "72148": 4, "73721": 3,
        # ER
        "99283": 5, "99284": 4,
        # Mental health
        "90837": 6, "90834": 8,
        # Preventive
        "99396": 10, "99397": 6,
        # Other
        "36415": 15, "93000": 5,
    }

    # Build weighted CPT list
    cpt_pool = []
    for cpt, weight in employer_cpts.items():
        cpt_pool.extend([cpt] * weight)

    employer_icd_map = {
        "99213": ["I10", "E11.9", "J06.9", "M54.5", "F32.9"],
        "99214": ["I10", "E11.9", "E78.5", "M17.11", "G43.909"],
        "99243": ["M17.11", "M54.5", "K21.0"],
        "99244": ["I25.10", "E11.65", "K80.20"],
        "27447": ["M17.11"],
        "66984": ["H25.11"],
        "43239": ["K21.0"],
        "23472": ["M75.111"],
        "80053": ["E11.9", "Z01.89"],
        "85025": ["D64.9", "Z01.89"],
        "82947": ["E11.9"],
        "83036": ["E11.9"],
        "71046": ["R05.9", "J18.9"],
        "72148": ["M54.5"],
        "73721": ["M17.11", "M23.91"],
        "99283": ["S93.401A", "R07.9"],
        "99284": ["R07.9", "S52.501A"],
        "90837": ["F32.9", "F41.1"],
        "90834": ["F32.9", "F41.1"],
        "99396": ["Z00.00"],
        "99397": ["Z00.00"],
        "36415": ["Z01.89"],
        "93000": ["I10", "R00.0"],
    }

    # Provider names for variety
    providers = [
        ("MARTINEZ", "ANNA", "1112223334"),
        ("KIM", "DAVID", "2223334445"),
        ("OCONNOR", "SEAN", "3334445556"),
        ("SHARMA", "PRIYA", "4445556667"),
        ("JOHNSON", "MARK", "5556667778"),
        ("BAYVIEW ORTHO GROUP", None, "6667778889"),
        ("CHESAPEAKE IMAGING LLC", None, "7778889990"),
        ("MD MENTAL HEALTH ASSOC", None, "8889990001"),
    ]

    claim_count = 0
    total_claims = 200

    for _ in range(total_claims):
        member = random.choice(members)
        cpt = random.choice(cpt_pool)
        service_date = _random_date_2024(1, 12)

        # Rates: in-network vs out-of-network
        is_oon = random.random() < 0.12  # 12% out of network
        medicare = MEDICARE_RATES.get(cpt, 100)

        if is_oon:
            billed = round(medicare * random.uniform(2.5, 5.0), 2)
            paid = round(medicare * random.uniform(0.8, 1.2), 2)  # Plan pays Medicare-ish
        else:
            multiplier = random.uniform(1.2, 1.8)
            billed = round(medicare * multiplier * 1.3, 2)
            paid = round(medicare * multiplier, 2)

        # High-cost outliers for surgical codes
        if cpt in ("27447", "66984", "43239", "23472"):
            facility_markup = random.uniform(2.0, 4.0) if random.random() < 0.4 else 1.0
            billed = round(billed * facility_markup, 2)
            paid = round(paid * facility_markup * 0.85, 2)

        # Some denials
        denied = random.random() < 0.08
        if denied:
            denial = random.choice(DENIAL_CODES)
            paid = 0
        else:
            denial = None

        adj_amt = round(billed - paid, 2)
        provider = random.choice(providers)
        claim_id = _random_claim_id()

        status = "1" if paid > 0 else "4"
        add(f"CLP*{claim_id}*{status}*{billed}*{paid}*{adj_amt}*MC*ECLAIM{claim_count:06d}~")
        add(f"NM1*QC*1*{member[0]}*{member[1]}****MI*{member[3]}~")

        if provider[1]:
            add(f"NM1*82*1*{provider[0]}*{provider[1]}****XX*{provider[2]}~")
        else:
            add(f"NM1*82*2*{provider[0]}*****XX*{provider[2]}~")

        add(f"SVC*HC:{cpt}*{billed}*{paid}**1~")
        add(f"DTM*472*{service_date}~")

        if denial:
            add(f"CAS*{denial[0]}*{denial[1]}*{adj_amt}~")
        elif adj_amt > 0:
            add(f"CAS*CO*45*{adj_amt}~")

        claim_count += 1

    # Trailer
    add(f"SE*{seg_count + 1}*0001~")
    add("GE*1*3~")
    add("IEA*1*000000003~")

    output = "\n".join(lines)
    path = os.path.join(OUTPUT_DIR, "maryland_manufacturing_835_2024.edi")
    with open(path, "w") as f:
        f.write(output)

    print(f"  [4] Employer 835: {path} ({claim_count} claims, {os.path.getsize(path):,} bytes)")


# ═══════════════════════════════════════════════════════════════════
# File 5: Employer SBC PDF
# ═══════════════════════════════════════════════════════════════════

def generate_sbc_pdf():
    """Generate SBC PDF for Maryland Precision Manufacturing PPO 2024."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    path = os.path.join(OUTPUT_DIR, "maryland_manufacturing_sbc_2024.pdf")
    doc = SimpleDocTemplate(path, pagesize=letter,
                            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                            topMargin=0.75 * inch, bottomMargin=0.75 * inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("SBCTitle", parent=styles["Heading1"],
                                 fontSize=16, textColor=colors.HexColor("#1a365d"),
                                 spaceAfter=6)
    subtitle_style = ParagraphStyle("SBCSub", parent=styles["Heading2"],
                                    fontSize=12, textColor=colors.HexColor("#2d3748"),
                                    spaceAfter=4)
    section_style = ParagraphStyle("SBCSection", parent=styles["Heading3"],
                                   fontSize=11, textColor=colors.white,
                                   backColor=colors.HexColor("#2b6cb0"),
                                   spaceAfter=4, spaceBefore=12,
                                   leftIndent=4, rightIndent=4,
                                   borderPadding=4)
    body_style = ParagraphStyle("SBCBody", parent=styles["Normal"],
                                fontSize=9, leading=12, spaceAfter=4)
    small_style = ParagraphStyle("SBCSmall", parent=styles["Normal"],
                                 fontSize=8, leading=10, textColor=colors.gray)

    story = []

    # Title block
    story.append(Paragraph("Summary of Benefits and Coverage", title_style))
    story.append(Paragraph(
        f"<b>{EMPLOYER_NAME}</b> — Maryland Precision Manufacturing PPO 2024", subtitle_style))
    story.append(Paragraph(
        "Coverage Period: 01/01/2024 - 12/31/2024 | Coverage for: Employee, Employee + Spouse, Employee + Children, Family",
        small_style))
    story.append(Spacer(1, 12))

    # Important questions table
    story.append(Paragraph("Important Questions", section_style))

    q_data = [
        ["Important Questions", "Answers", "Why This Matters"],
        ["What is the overall\ndeductible?",
         "$1,500 individual\n$3,000 family",
         "Generally, you must pay all of the costs from providers up to the deductible amount before this plan begins to pay. Copays don't count toward the deductible."],
        ["Are there services\ncovered before you meet\nyour deductible?",
         "Yes. Preventive care,\nprimary care copays,\nand generic drugs.",
         "This plan covers some items and services even if you haven't yet met the deductible amount."],
        ["Are there other\ndeductibles for specific\nservices?",
         "Yes. $3,000 for\nout-of-network services.",
         "You must pay all of the costs for out-of-network services up to the separate deductible before the plan pays."],
        ["What is the out-of-pocket\nlimit for this plan?",
         "$5,000 individual\n$10,000 family",
         "The out-of-pocket limit is the most you could pay in a year for covered services."],
        ["What is not included in\nthe out-of-pocket limit?",
         "Premiums, balance-billing\ncharges, penalties for\nnon-compliance.",
         "Even though you pay these expenses, they don't count toward the out-of-pocket limit."],
        ["Will you pay less if you\nuse a network provider?",
         "Yes. In-network: 20%\ncoinsurance.\nOut-of-network: 40%.",
         "This plan uses a provider network. You will pay less if you use a provider in the plan's network."],
        ["Do you need a referral\nto see a specialist?",
         "No.",
         "You can see the specialist you choose without a referral."],
    ]

    q_table = Table(q_data, colWidths=[1.8 * inch, 1.8 * inch, 3.2 * inch])
    q_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2b6cb0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
    ]))
    story.append(q_table)
    story.append(Spacer(1, 16))

    # Cost sharing table
    story.append(Paragraph("Common Medical Events — Cost Sharing", section_style))

    cost_data = [
        ["Common Medical Event", "Services You May Need", "In-Network\n(You Pay)", "Out-of-Network\n(You Pay)"],
        ["If you visit a health\ncare provider's office\nor clinic",
         "Primary care visit\nSpecialist visit\nPreventive care",
         "$30 copay/visit\n$60 copay/visit\nNo charge",
         "40% coinsurance\n40% coinsurance\n40% coinsurance"],
        ["If you have a test",
         "Diagnostic test (x-ray, blood work)\nImaging (CT/PET/MRI)",
         "20% coinsurance\n20% coinsurance",
         "40% coinsurance\n40% coinsurance"],
        ["If you need drugs to\ntreat your illness or\ncondition",
         "Generic drugs (Tier 1)\nPreferred brand (Tier 2)\nNon-preferred brand (Tier 3)\nSpecialty drugs (Tier 4)",
         "$15 copay\n$45 copay\n$75 copay\n30% coinsurance",
         "$30 copay\n$90 copay\nNot covered\nNot covered"],
        ["If you have outpatient\nsurgery",
         "Facility fee\nPhysician/surgeon fees",
         "20% coinsurance\n20% coinsurance",
         "40% coinsurance\n40% coinsurance"],
        ["If you need immediate\nmedical attention",
         "Emergency room care\nUrgent care",
         "$300 copay/visit\n$60 copay/visit",
         "$300 copay/visit\n40% coinsurance"],
        ["If you have a hospital\nstay",
         "Facility fee\nPhysician/surgeon fees",
         "20% coinsurance\n20% coinsurance",
         "40% coinsurance\n40% coinsurance"],
        ["If you need mental\nhealth or substance\nabuse services",
         "Outpatient services\nInpatient services",
         "$30 copay/visit\n20% coinsurance",
         "40% coinsurance\n40% coinsurance"],
        ["If you are pregnant",
         "Office visits\nChildbirth/delivery\nprofessional services\nChildbirth/delivery\nfacility services",
         "$30 copay/visit\n20% coinsurance\n\n20% coinsurance",
         "40% coinsurance\n40% coinsurance\n\n40% coinsurance"],
    ]

    c_table = Table(cost_data, colWidths=[1.7 * inch, 2.0 * inch, 1.5 * inch, 1.5 * inch])
    c_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2b6cb0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
    ]))
    story.append(c_table)
    story.append(Spacer(1, 16))

    # Excluded services
    story.append(Paragraph("Excluded Services & Other Covered Services", section_style))
    story.append(Paragraph(
        "<b>Services Your Plan Generally Does NOT Cover:</b> Cosmetic surgery, dental care (adult), "
        "long-term care, non-emergency care when traveling outside the U.S., routine foot care, "
        "weight loss programs, private-duty nursing, infertility treatment.",
        body_style))
    story.append(Paragraph(
        "<b>Other Covered Services (Limitations may apply):</b> Acupuncture (20 visits/year), "
        "bariatric surgery (prior authorization required), chiropractic care (30 visits/year), "
        "hearing aids (1 per ear every 3 years), home health care (60 visits/year), "
        "hospice services, routine eye care (1 exam/year).",
        body_style))
    story.append(Spacer(1, 12))

    # Footer
    story.append(Paragraph(
        "Questions: Call Member Services at 1-800-555-0199 or visit www.mpmhealthplan.example.com. "
        "If you aren't clear about any of the underlined terms used in this form, see the Glossary.",
        small_style))
    story.append(Paragraph(
        f"Plan Year: January 1, 2024 - December 31, 2024 | Plan Sponsor: {EMPLOYER_NAME} | "
        "EIN: 52-1234567 | Plan Number: 501",
        small_style))
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        "This is a summary of the most important features of this health coverage. "
        "This is not the complete terms of coverage. For complete details, see the plan document. "
        "This document is for illustrative/testing purposes only.",
        ParagraphStyle("Disclaimer", parent=small_style, textColor=colors.HexColor("#e53e3e"))))

    doc.build(story)
    print(f"  [5] Employer SBC PDF: {path} ({os.path.getsize(path):,} bytes)")


# ═══════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════

def main():
    print(f"\nGenerating CivicScale test files in {OUTPUT_DIR}/\n")

    # File 1: Provider 835
    claim_data = generate_provider_835()

    # File 2: Provider 837P (uses same claims)
    generate_provider_837(claim_data)

    # File 3: Contract fee schedule Excel
    generate_contract_excel()

    # File 4: Employer 835
    generate_employer_835()

    # File 5: Employer SBC PDF
    generate_sbc_pdf()

    print(f"\nAll 5 files generated successfully in {OUTPUT_DIR}/")


if __name__ == "__main__":
    main()
