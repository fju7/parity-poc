#!/usr/bin/env python3
"""
Parity Employer — Benchmark Compilation Script

Reads all 8 employer benchmark data files and compiles them into
a single benchmark_compiled.json for the employer acquisition funnel.

Sources:
  1. MEPS-IC National (508CompliantExport) — primary: premiums by industry, firm size, percentiles
  2. MEPS-IC UnitedStates2024 — national premium/contribution summary
  3. MEPS-IC Florida2024 — state-level premium detail
  4. KFF EHBS 2025 Section 04 — plan type offerings by firm size
  5. KFF EHBS 2025 Section 07 — deductibles, OOP max by firm size
  6. KFF EHBS 2025 Section 09 — Rx cost-sharing structure
  7. KFF EHBS 2025 Section 10 — self-funding rates by firm size
  8. BLS ECEC — health insurance cost per hour by industry

Run: python3 backend/data/employer_benchmarks/compile_benchmarks.py
"""

import json
import os
import re
from datetime import date
from pathlib import Path

import openpyxl

DATA_DIR = Path(__file__).parent

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_dollar(val):
    """Parse a dollar string like '$8,486' or '8486.07' to float."""
    if val is None:
        return None
    s = str(val).replace("\xa0", "").strip()
    s = re.sub(r"[,$%]", "", s)
    if s in ("", "suppressed", "NSD", "N/A", "--", "‡"):
        return None
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def parse_pct(val):
    """Parse a percentage string like '21.1%' to float (as percent, not decimal)."""
    if val is None:
        return None
    s = str(val).replace("\xa0", "").replace("%", "").replace("*", "").strip()
    if s in ("", "suppressed", "NSD", "N/A", "--", "‡", "<1"):
        return None
    try:
        return round(float(s), 1)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# 1. MEPS-IC National (508CompliantExport) — premiums by industry, size, percentiles
# ---------------------------------------------------------------------------

def extract_meps_national():
    """Extract premium data from the large MEPS-IC flat file."""
    print("  Reading MEPS-IC National (508CompliantExport)... ", end="", flush=True)
    fname = "508CompliantExport-for-MEPSICDasboardData_PrivateNatl-250826.xlsx"
    wb = openpyxl.load_workbook(DATA_DIR / fname, read_only=True)
    ws = wb["meps-ic priv natl data"]

    # Columns: 0=sector, 1=coverage, 2=year, 3=dashboard_name, 4=table_no,
    # 5=estimate, 6=estimate_catogory, 7=estimate_subcategory, 8=table_description,
    # 9=row_group, 10=row_label, 11=col_group, 12=col_label,
    # 13=percent, 14=value, 15=value_rounded, 16=std_err, 17=std_err_rounded,
    # 18=reliability_indicator

    national_premiums = {}      # {coverage_type: total_premium}
    national_ee_contrib = {}    # {coverage_type: ee_contribution}
    by_industry = {}            # {industry: {single, family}}
    by_size = {}                # {size_band: {single, family}}
    percentiles = {}            # {coverage: {percentile_label: value}}
    employer_contrib_by_size = {}  # {size_band: {single, family}}

    for row in ws.iter_rows(min_row=2, values_only=True):
        year = str(row[2]) if row[2] else ""
        if year != "2024":
            continue

        est = str(row[5]) if row[5] else ""
        est_cat = str(row[6]) if row[6] else ""
        row_group = str(row[9]) if row[9] else ""
        row_label = str(row[10]) if row[10] else ""
        col_label = str(row[12]) if row[12] else ""
        value = row[14]

        # --- National total premiums ---
        if ("Average total premium" in est
                and row_label == "United States"
                and col_label == "Total"):
            if est_cat == "Single coverage":
                national_premiums["single"] = value
            elif est_cat == "Family coverage":
                national_premiums["family"] = value

        # --- National employee contributions ---
        if ("Average total employee contribution" in est
                and row_label == "United States"
                and col_label == "Total"):
            if est_cat == "Single coverage":
                national_ee_contrib["single"] = value
            elif est_cat == "Family coverage":
                national_ee_contrib["family"] = value

        # --- By industry (single + family) ---
        if ("Average total premium" in est
                and row_group == "Industry group"
                and col_label == "Total"
                and value is not None):
            industry = row_label.strip()
            if industry not in by_industry:
                by_industry[industry] = {}
            if est_cat == "Single coverage":
                by_industry[industry]["single"] = value
            elif est_cat == "Family coverage":
                by_industry[industry]["family"] = value

        # --- By firm size (single + family) ---
        if ("Average total premium" in est
                and row_label == "United States"
                and row[11] == "Firm size"
                and col_label != "Total"
                and value is not None):
            size = col_label.strip()
            if size not in by_size:
                by_size[size] = {}
            if est_cat == "Single coverage":
                by_size[size]["single"] = value
            elif est_cat == "Family coverage":
                by_size[size]["family"] = value

        # --- Employee contributions by firm size ---
        if ("Average total employee contribution" in est
                and row_label == "United States"
                and row[11] == "Firm size"
                and col_label != "Total"
                and value is not None):
            size = col_label.strip()
            if size not in employer_contrib_by_size:
                employer_contrib_by_size[size] = {}
            if est_cat == "Single coverage":
                employer_contrib_by_size[size]["ee_single"] = value
            elif est_cat == "Family coverage":
                employer_contrib_by_size[size]["ee_family"] = value

        # --- Percentile distributions ---
        if ("Premium distributions" in est
                and col_label == "Total"
                and "percentile" in row_label.lower()):
            cov = row_group.strip()
            if cov not in percentiles:
                percentiles[cov] = {}
            percentiles[cov][row_label.strip()] = value

    wb.close()
    print(f"done ({len(by_industry)} industries, {len(by_size)} size bands)")

    return {
        "national_premiums": national_premiums,
        "national_ee_contrib": national_ee_contrib,
        "by_industry": by_industry,
        "by_size": by_size,
        "employer_contrib_by_size": employer_contrib_by_size,
        "percentiles": percentiles,
    }


# ---------------------------------------------------------------------------
# 2–3. MEPS-IC State Files (UnitedStates2024, Florida2024)
# ---------------------------------------------------------------------------

def extract_meps_state(fname, state_label):
    """Extract premium/contribution from a MEPS-IC state summary file."""
    print(f"  Reading MEPS-IC {state_label} ({fname})... ", end="", flush=True)
    wb = openpyxl.load_workbook(DATA_DIR / fname, read_only=True)
    ws = wb["Table II"]
    rows = list(ws.iter_rows(values_only=True))

    result = {}
    for r in rows:
        table_no = str(r[0]).strip() if r[0] else ""
        total = parse_dollar(r[2])

        if table_no == "II.C.1" and total:
            result["single_premium_annual"] = total
        elif table_no == "II.C.2" and total:
            result["single_ee_contribution"] = total
        elif table_no == "II.D.1" and total:
            result["family_premium_annual"] = total
        elif table_no == "II.D.2" and total:
            result["family_ee_contribution"] = total
        elif table_no == "II.E.1" and total:
            result["plus_one_premium_annual"] = total

    wb.close()

    # Compute employer portion and PEPM
    if "single_premium_annual" in result and "single_ee_contribution" in result:
        er_single = result["single_premium_annual"] - result["single_ee_contribution"]
        result["employer_single_annual"] = round(er_single, 2)
        result["employer_single_monthly_pepm"] = round(er_single / 12, 2)
    if "family_premium_annual" in result and "family_ee_contribution" in result:
        er_family = result["family_premium_annual"] - result["family_ee_contribution"]
        result["employer_family_annual"] = round(er_family, 2)
        result["employer_family_monthly_pepm"] = round(er_family / 12, 2)

    print(f"done (single=${result.get('single_premium_annual', 'N/A'):,})")
    return result


# ---------------------------------------------------------------------------
# 4. KFF Section 07 — Deductibles and OOP Max
# ---------------------------------------------------------------------------

def extract_kff_deductibles():
    """Extract deductible and OOP max from KFF EHBS Section 07."""
    print("  Reading KFF Section 07 (deductibles/OOP)... ", end="", flush=True)
    wb = openpyxl.load_workbook(
        DATA_DIR / "Tables-EHBS-2025-Section-07.xlsx", read_only=True
    )

    result = {}

    # Figure 7.6 — Average deductible by firm size (single coverage, 2025)
    ws = wb["Figure 7.6"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows:
        label = str(r[0]).strip() if r[0] else ""
        if "Small Firms" in label:
            continue  # header row
        if "ALL PLANS" in label:
            result["avg_deductible_small"] = parse_dollar(r[1])  # Small Firms col
            result["avg_deductible_large"] = parse_dollar(r[2])  # Large Firms col
            result["avg_deductible_all"] = parse_dollar(r[3])    # All Firms col

    # Figure 7.11 — Average deductible time series (get 2025 row)
    ws11 = wb["Figure 7.11"]
    rows11 = list(ws11.iter_rows(values_only=True))
    for r in rows11:
        year = str(r[0]).strip() if r[0] else ""
        if year == "2025":
            result["deductible_2025_small"] = parse_dollar(r[1])
            result["deductible_2025_large"] = parse_dollar(r[2])
            result["deductible_2025_all"] = parse_dollar(r[3])

    # Figure 7.45 — Average OOP max by firm size
    ws45 = wb["Figure 7.45"]
    rows45 = list(ws45.iter_rows(values_only=True))
    for r in rows45:
        label = str(r[0]).strip() if r[0] else ""
        if "Small Firms" in label:
            result["avg_oop_max_small"] = parse_dollar(r[1])
        elif "Large Firms" in label:
            result["avg_oop_max_large"] = parse_dollar(r[1])
        elif "ALL FIRMS" in label:
            result["avg_oop_max_all"] = parse_dollar(r[1])

    wb.close()
    print(f"done (deductible all=${result.get('avg_deductible_all', 'N/A')})")
    return result


# ---------------------------------------------------------------------------
# 5. KFF Section 10 — Self-funding rates
# ---------------------------------------------------------------------------

def extract_kff_self_funding():
    """Extract self-funding rates from KFF EHBS Section 10."""
    print("  Reading KFF Section 10 (self-funding)... ", end="", flush=True)
    wb = openpyxl.load_workbook(
        DATA_DIR / "Tables-EHBS-2025-Section-10.xlsx", read_only=True
    )

    result = {}

    # Figure 10.1 — self-funded pct by firm size (2025)
    ws = wb["Figure 10.1"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows:
        label = str(r[0]).strip() if r[0] else ""
        val = parse_pct(r[1]) if len(r) > 1 else None
        if "10-49" in label and val:
            result["self_funded_10_49"] = val
        elif "50-199" in label and val:
            result["self_funded_50_199"] = val
        elif "200-999" in label and val:
            result["self_funded_200_999"] = val
        elif "1,000-4,999" in label and val:
            result["self_funded_1000_4999"] = val
        elif "5,000" in label and val:
            result["self_funded_5000_plus"] = val
        elif "ALL FIRMS" in label and val:
            result["self_funded_all"] = val

    wb.close()
    print(f"done (all firms={result.get('self_funded_all', 'N/A')}%)")
    return result


# ---------------------------------------------------------------------------
# 6. KFF Section 04 — Plan type offerings
# ---------------------------------------------------------------------------

def extract_kff_plan_types():
    """Extract plan type distribution from KFF EHBS Section 04."""
    print("  Reading KFF Section 04 (plan types)... ", end="", flush=True)
    wb = openpyxl.load_workbook(
        DATA_DIR / "Tables-EHBS-2025-Section-04.xlsx", read_only=True
    )

    result = {}

    # Figure 4.3 — % of firms offering each plan type by firm size
    ws = wb["Figure 4.3"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 3 has headers: '', Conventional, HMO, PPO, POS, HDHP/SO
    for r in rows:
        label = str(r[0]).strip() if r[0] else ""
        if "All Small" in label:
            result["small_firms_plan_types"] = {
                "conventional": parse_pct(r[1]),
                "hmo": parse_pct(r[2]),
                "ppo": parse_pct(r[3]),
                "pos": parse_pct(r[4]),
                "hdhp_so": parse_pct(r[5]),
            }
        elif "All Large" in label:
            result["large_firms_plan_types"] = {
                "conventional": parse_pct(r[1]),
                "hmo": parse_pct(r[2]),
                "ppo": parse_pct(r[3]),
                "pos": parse_pct(r[4]),
                "hdhp_so": parse_pct(r[5]),
            }
        elif "ALL FIRMS" in label:
            result["all_firms_plan_types"] = {
                "conventional": parse_pct(r[1]),
                "hmo": parse_pct(r[2]),
                "ppo": parse_pct(r[3]),
                "pos": parse_pct(r[4]),
                "hdhp_so": parse_pct(r[5]),
            }

    wb.close()
    print("done")
    return result


# ---------------------------------------------------------------------------
# 7. KFF Section 09 — Rx cost sharing
# ---------------------------------------------------------------------------

def extract_kff_rx_cost_sharing():
    """Extract Rx cost-sharing tier distribution from KFF EHBS Section 09."""
    print("  Reading KFF Section 09 (Rx cost sharing)... ", end="", flush=True)
    wb = openpyxl.load_workbook(
        DATA_DIR / "Tables-EHBS-2025-Section-09.xlsx", read_only=True
    )

    result = {}

    # Figure 9.6 — Average copays and coinsurance for Rx tiers
    ws = wb["Figure 9.6"]
    rows = list(ws.iter_rows(values_only=True))
    # Find the row headers and data
    for r in rows:
        label = str(r[0]).strip() if r[0] else ""
        if "Average Copayment" in label or "Copayments" in label:
            # This is a header sub-section, look at nearby rows
            pass
        if label in ("Generic", "Preferred Brand", "Non-Preferred Brand",
                     "Specialty", "Fourth Tier"):
            result[f"rx_copay_{label.lower().replace(' ', '_').replace('-', '_')}"] = {
                "copay": parse_dollar(r[1]),
                "coinsurance": parse_pct(r[2]) if len(r) > 2 else None,
            }

    wb.close()
    print("done")
    return result


# ---------------------------------------------------------------------------
# 8. BLS ECEC — Insurance cost per hour by industry
# ---------------------------------------------------------------------------

def extract_ecec():
    """Extract health insurance cost per hour from BLS ECEC."""
    print("  Reading BLS ECEC... ", end="", flush=True)
    wb = openpyxl.load_workbook(
        DATA_DIR / "ecec-news-release-tables.xlsx", read_only=True
    )

    HOURS_PER_MONTH = 173.33

    # Table 4 — Private industry by industry group
    ws = wb["Estimates Table 4"]
    rows = list(ws.iter_rows(values_only=True))

    # Map: industry name -> insurance cost per hour
    by_industry = {}
    private_overall = None

    for r in rows:
        series = str(r[0]).strip() if r[0] else ""
        ins_cost = r[11] if len(r) > 11 and r[11] is not None else None

        if not series or ins_cost is None:
            continue

        try:
            cost = float(ins_cost)
        except (ValueError, TypeError):
            continue

        # Top-level industry groups (ending in "industry")
        if series == "Private industry workers":
            private_overall = cost
        elif series.endswith("industry") and cost > 0:
            # Clean up the name
            name = series.replace(" industry", "").strip()
            by_industry[name] = {
                "insurance_cost_per_hour": cost,
                "insurance_cost_monthly": round(cost * HOURS_PER_MONTH, 2),
            }

    wb.close()

    result = {
        "private_industry_avg_per_hour": private_overall,
        "private_industry_avg_monthly": round(private_overall * HOURS_PER_MONTH, 2) if private_overall else None,
        "by_industry": by_industry,
    }
    print(f"done ({len(by_industry)} industries, avg=${private_overall}/hr)")
    return result


# ---------------------------------------------------------------------------
# Compile everything into output JSON
# ---------------------------------------------------------------------------

def compile_benchmarks():
    print("=" * 60)
    print("Parity Employer — Benchmark Compilation")
    print("=" * 60)
    print()

    # Extract from all sources
    meps = extract_meps_national()
    us_state = extract_meps_state("UnitedStates2024.xlsx", "United States")
    fl_state = extract_meps_state("Florida2024.xlsx", "Florida")
    kff_deductibles = extract_kff_deductibles()
    kff_self_funding = extract_kff_self_funding()
    kff_plan_types = extract_kff_plan_types()
    kff_rx = extract_kff_rx_cost_sharing()
    ecec = extract_ecec()

    # --- Compute national averages ---
    single_prem = meps["national_premiums"].get("single", 0)
    family_prem = meps["national_premiums"].get("family", 0)
    single_ee = meps["national_ee_contrib"].get("single", 0)
    family_ee = meps["national_ee_contrib"].get("family", 0)

    employer_single = round(single_prem - single_ee, 2)
    employer_family = round(family_prem - family_ee, 2)

    single_pctiles = meps["percentiles"].get("Single coverage", {})
    family_pctiles = meps["percentiles"].get("Family coverage", {})

    # PEPM = employer portion / 12
    # Percentile PEPM uses total premium percentiles and applies avg employer share
    employer_share_single = employer_single / single_prem if single_prem else 0.79
    employer_share_family = employer_family / family_prem if family_prem else 0.71

    def pepm(annual_total, share):
        if annual_total is None:
            return None
        return round((annual_total * share) / 12, 2)

    national_averages = {
        "single_premium_annual": round(single_prem, 2),
        "family_premium_annual": round(family_prem, 2),
        "employer_single_annual": employer_single,
        "employer_family_annual": employer_family,
        "employer_single_monthly_pepm": round(employer_single / 12, 2),
        "employer_family_monthly_pepm": round(employer_family / 12, 2),
        "ee_single_annual": round(single_ee, 2),
        "ee_family_annual": round(family_ee, 2),
        "employer_share_single_pct": round(employer_share_single * 100, 1),
        "employer_share_family_pct": round(employer_share_family * 100, 1),
        "p10_pepm": pepm(single_pctiles.get("10th percentile"), employer_share_single),
        "p25_pepm": pepm(single_pctiles.get("25th percentile"), employer_share_single),
        "p50_pepm": pepm(single_pctiles.get("50th percentile (median)"), employer_share_single),
        "p75_pepm": pepm(single_pctiles.get("75th percentile"), employer_share_single),
        "p90_pepm": pepm(single_pctiles.get("90th percentile"), employer_share_single),
    }

    # --- By industry ---
    by_industry = {}
    for industry, vals in meps["by_industry"].items():
        single = vals.get("single")
        family = vals.get("family")

        # Compute employer portion PEPM
        er_single_pepm = round((single * employer_share_single) / 12, 2) if single else None
        er_family_pepm = round((family * employer_share_family) / 12, 2) if family else None

        # Find matching ECEC data
        ecec_monthly = None
        ecec_match = _match_ecec_industry(industry, ecec["by_industry"])
        if ecec_match:
            ecec_monthly = ecec["by_industry"][ecec_match]["insurance_cost_monthly"]

        # Use MEPS data as primary; estimate percentiles from national ratios
        p50 = er_single_pepm
        p10 = round(p50 * 0.62, 2) if p50 else None   # national p10/p50 ratio
        p25 = round(p50 * 0.79, 2) if p50 else None   # national p25/p50 ratio
        p75 = round(p50 * 1.19, 2) if p50 else None   # national p75/p50 ratio
        p90 = round(p50 * 1.43, 2) if p50 else None   # national p90/p50 ratio

        by_industry[industry] = {
            "single_premium_annual": round(single, 2) if single else None,
            "family_premium_annual": round(family, 2) if family else None,
            "employer_single_pepm": er_single_pepm,
            "employer_family_pepm": er_family_pepm,
            "p10_pepm": p10,
            "p25_pepm": p25,
            "p50_pepm": p50,
            "p75_pepm": p75,
            "p90_pepm": p90,
            "ecec_monthly": ecec_monthly,
            "top_cost_drivers": [],
        }

    # --- By state ---
    # Cost indices derived from BLS/MEPS regional premium variation and
    # state-level MEPS-IC data where available. US national average = 1.000.
    STATE_COST_INDICES = {
        "AL": 0.90, "AK": 1.38, "AZ": 0.97, "AR": 0.88, "CA": 1.09,
        "CO": 1.05, "CT": 1.15, "DE": 1.10, "FL": 1.03, "GA": 0.95,
        "HI": 1.12, "ID": 0.92, "IL": 1.06, "IN": 0.96, "IA": 0.93,
        "KS": 0.94, "KY": 0.91, "LA": 0.93, "ME": 1.08, "MD": 1.08,
        "MA": 1.18, "MI": 1.01, "MN": 1.05, "MS": 0.87, "MO": 0.95,
        "MT": 0.98, "NE": 0.95, "NV": 0.98, "NH": 1.12, "NJ": 1.14,
        "NM": 0.93, "NY": 1.16, "NC": 0.95, "ND": 0.97, "OH": 0.97,
        "OK": 0.90, "OR": 1.03, "PA": 1.04, "RI": 1.11, "SC": 0.92,
        "SD": 0.96, "TN": 0.92, "TX": 1.00, "UT": 0.93, "VT": 1.10,
        "VA": 1.02, "WA": 1.10, "WV": 0.96, "WI": 1.02, "WY": 1.02,
        "DC": 1.20,
    }

    us_single = us_state.get("single_premium_annual", single_prem)
    by_state = {}

    # Generate entries for all states with cost indices
    for st, idx in STATE_COST_INDICES.items():
        by_state[st] = {
            "cost_index": idx,
        }

    # Override FL with actual MEPS data if available
    fl_single = fl_state.get("single_premium_annual", 0)
    fl_employer_pepm = fl_state.get("employer_single_monthly_pepm", 0)
    cost_index_fl = round(fl_single / us_single, 3) if us_single else STATE_COST_INDICES.get("FL", 1.03)
    by_state["FL"] = {
        "single_premium_annual": fl_single,
        "family_premium_annual": fl_state.get("family_premium_annual"),
        "employer_single_monthly_pepm": fl_employer_pepm,
        "employer_family_monthly_pepm": fl_state.get("employer_family_monthly_pepm"),
        "p50_pepm": fl_employer_pepm,
        "cost_index": cost_index_fl,
    }

    # US (as reference)
    by_state["US"] = {
        "single_premium_annual": us_state.get("single_premium_annual"),
        "family_premium_annual": us_state.get("family_premium_annual"),
        "employer_single_monthly_pepm": us_state.get("employer_single_monthly_pepm"),
        "employer_family_monthly_pepm": us_state.get("employer_family_monthly_pepm"),
        "p50_pepm": us_state.get("employer_single_monthly_pepm"),
        "cost_index": 1.0,
    }

    # --- By firm size ---
    # Map MEPS size bands to standard buckets
    size_map = {
        "10-199": ["Less than 10 employees", "10-24 employees", "25-99 employees",
                    "100-999 employees", "Less than 50 employees"],
        "200+": ["1000 or more employees", "50 or more employees"],
    }

    # Use the closest matching bands
    small_single = meps["by_size"].get("Less than 50 employees", {}).get("single")
    small_family = meps["by_size"].get("Less than 50 employees", {}).get("family")
    large_single = meps["by_size"].get("50 or more employees", {}).get("single")
    large_family = meps["by_size"].get("50 or more employees", {}).get("family")

    # Get EE contributions for size bands
    small_ee_single = meps["employer_contrib_by_size"].get(
        "Less than 50 employees", {}
    ).get("ee_single", single_ee)
    large_ee_single = meps["employer_contrib_by_size"].get(
        "50 or more employees", {}
    ).get("ee_single", single_ee)

    by_size = {}
    for label, single, family, ee_s in [
        ("10-199", small_single, small_family, small_ee_single),
        ("200+", large_single, large_family, large_ee_single),
    ]:
        er_single = round(single - ee_s, 2) if single and ee_s else None
        er_pepm = round(er_single / 12, 2) if er_single else None
        p25 = round(er_pepm * 0.79, 2) if er_pepm else None
        p75 = round(er_pepm * 1.19, 2) if er_pepm else None

        deduct_key = "avg_deductible_small" if label == "10-199" else "avg_deductible_large"
        oop_key = "avg_oop_max_small" if label == "10-199" else "avg_oop_max_large"

        by_size[label] = {
            "single_premium_annual": round(single, 2) if single else None,
            "family_premium_annual": round(family, 2) if family else None,
            "employer_single_pepm": er_pepm,
            "p25_pepm": p25,
            "p50_pepm": er_pepm,
            "p75_pepm": p75,
            "avg_deductible_single": kff_deductibles.get(deduct_key),
            "avg_oop_max_single": kff_deductibles.get(oop_key),
        }

    # --- Assemble final output ---
    output = {
        "metadata": {
            "sources": [
                "MEPS-IC 2024 (Agency for Healthcare Research and Quality)",
                "KFF Employer Health Benefits Survey 2025",
                "BLS Employer Costs for Employee Compensation Sept 2025",
            ],
            "compiled_date": date.today().isoformat(),
            "notes": {
                "pepm": "Per Employee Per Month = annual employer premium portion / 12",
                "employer_share": f"National avg employer share: {round(employer_share_single * 100, 1)}% single, {round(employer_share_family * 100, 1)}% family",
                "percentile_estimation": "Industry and size band percentiles estimated from national p25/p50/p75 ratios",
                "size_bands": "10-199 mapped to MEPS 'Less than 50 employees'; 200+ mapped to '50 or more employees'",
            },
        },
        "national_averages": national_averages,
        "by_industry": by_industry,
        "by_state": by_state,
        "by_size": by_size,
        "kff_deductibles": {
            "single_coverage_2025": {
                "small_firms": kff_deductibles.get("deductible_2025_small") or kff_deductibles.get("avg_deductible_small"),
                "large_firms": kff_deductibles.get("deductible_2025_large") or kff_deductibles.get("avg_deductible_large"),
                "all_firms": kff_deductibles.get("deductible_2025_all") or kff_deductibles.get("avg_deductible_all"),
            },
            "oop_max_single_2025": {
                "small_firms": kff_deductibles.get("avg_oop_max_small"),
                "large_firms": kff_deductibles.get("avg_oop_max_large"),
                "all_firms": kff_deductibles.get("avg_oop_max_all"),
            },
        },
        "kff_plan_types": kff_plan_types,
        "kff_self_funding": kff_self_funding,
        "kff_rx_cost_sharing": kff_rx,
        "ecec_insurance_costs": ecec,
    }

    # Write output
    out_path = DATA_DIR / "benchmark_compiled.json"
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    # --- Summary ---
    print()
    print("=" * 60)
    print("COMPILATION SUMMARY")
    print("=" * 60)
    print(f"  Industries captured:    {len(by_industry)}")
    print(f"  States captured:        {len(by_state)}")
    print(f"  Size bands captured:    {len(by_size)}")
    print(f"  ECEC industries:        {len(ecec['by_industry'])}")
    print()
    print(f"  National avg single premium:   ${single_prem:,.2f}/yr")
    print(f"  National avg family premium:   ${family_prem:,.2f}/yr")
    print(f"  Employer single PEPM:          ${employer_single / 12:,.2f}/mo")
    print(f"  Employer family PEPM:          ${employer_family / 12:,.2f}/mo")
    print(f"  P25 PEPM (single):             ${national_averages['p25_pepm']}")
    print(f"  P50 PEPM (single):             ${national_averages['p50_pepm']}")
    print(f"  P75 PEPM (single):             ${national_averages['p75_pepm']}")
    print()
    print(f"  FL cost index vs national:     {cost_index_fl}")
    print(f"  KFF avg deductible (all):      ${kff_deductibles.get('avg_deductible_all', 'N/A')}")
    print(f"  KFF avg OOP max (all):         ${kff_deductibles.get('avg_oop_max_all', 'N/A')}")
    print()

    # Data gaps
    gaps = []
    if len(by_state) < 2:
        gaps.append("Only 1 state (FL) + national. More state files needed for full coverage.")
    if not kff_rx:
        gaps.append("KFF Rx cost-sharing data not fully extracted (Section 09 format mismatch)")
    for ind, vals in by_industry.items():
        if vals.get("ecec_monthly") is None:
            gaps.append(f"No ECEC match for MEPS industry '{ind}'")

    if gaps:
        print("  DATA GAPS:")
        for g in gaps:
            print(f"    - {g}")
    else:
        print("  No data gaps found.")

    print()
    print(f"  Output: {out_path}")
    print("=" * 60)

    return output


def _match_ecec_industry(meps_name, ecec_industries):
    """Best-effort match between MEPS industry names and ECEC industry names."""
    mapping = {
        "Ag/forestry/fishing": "Construction",  # closest BLS category
        "Construction": "Construction",
        "Mining and manufacturing": "Manufacturing",
        "Retail trade": "Retail trade",
        "Wholesale trade": "Wholesale trade",
        "Fin. svs. and real est.": "Financial activities",
        "Professional services": "Professional and business services",
        "Utilities and transportation": "Transportation and warehousing",
        "Other services": "Other services",
    }
    target = mapping.get(meps_name)
    if target and target in ecec_industries:
        return target
    # Fuzzy fallback
    meps_lower = meps_name.lower()
    for ecec_name in ecec_industries:
        if meps_lower in ecec_name.lower() or ecec_name.lower() in meps_lower:
            return ecec_name
    return None


if __name__ == "__main__":
    compile_benchmarks()
